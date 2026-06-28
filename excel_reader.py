"""Reads recipient data from uploaded Excel workbooks (in-memory)."""
import openpyxl

REQUIRED_COLS = ["Email", "Name", "Enrollment Month", "Enrollment ending Month", "DA"]
HEADER_ROW = 1                 # row with column titles
DATA_START = HEADER_ROW + 1    # scan every row after the header; rows without
                               # a valid email (blanks/spacers) are skipped anyway


def read_workbook(file_or_path):
    """
    Accepts a file path OR a file-like object (e.g. an upload stream).
    Returns a list of {"sheet_name": str, "recipients": [ {col: val}, ... ]}.
    Rows without a valid email are skipped. Extra columns are ignored.
    """
    try:
        wb = openpyxl.load_workbook(file_or_path, data_only=True)
    except Exception as e:
        raise RuntimeError(f"Cannot open workbook: {e}")

    sheets = []
    for ws in wb.worksheets:
        col_index = {}
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=HEADER_ROW, column=col).value
            if cell_val is None:
                continue
            header = str(cell_val).strip()
            if header in REQUIRED_COLS:
                col_index[header] = col

        if "Email" not in col_index:
            continue

        recipients = []
        for row_idx in range(DATA_START, ws.max_row + 1):
            email_cell = ws.cell(row=row_idx, column=col_index["Email"]).value
            if not email_cell:
                continue
            email = str(email_cell).strip()
            if "@" not in email:
                continue

            recipient = {"Email": email}
            for col_name in REQUIRED_COLS:
                if col_name == "Email":
                    continue
                if col_name in col_index:
                    raw = ws.cell(row=row_idx, column=col_index[col_name]).value
                    recipient[col_name] = str(raw).strip() if raw is not None else ""
                else:
                    recipient[col_name] = ""
            recipients.append(recipient)

        sheets.append({"sheet_name": ws.title, "recipients": recipients})

    wb.close()
    return sheets
